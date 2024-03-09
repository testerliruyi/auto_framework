"""
func:读取各类文件内容
"""
import yaml
import openpyxl
from common.file_handle.config_parser import MyConfigParser


class ReadFile:

    # 读取文本文件
    @staticmethod
    def read_txt_file(filename) -> list[str]:
        with open(filename, 'r', encoding='utf-8') as file:
            content_list = file.readlines()
        return content_list

    # 读取配置文件
    @staticmethod
    def read_config_file(path, selector=None, option=None):
        conf = MyConfigParser()
        conf.read(path)
        if not option:
            # 加果option为空，则返回内所有配量项
            result = dict(conf.items(selector))
        elif not selector:
            result = dict(conf.keys())
        else:
            # option不为空时，则返回对应配置 options
            result = conf.get(selector, option)
        return result

    # 读取yaml文件
    @staticmethod
    def read_yaml_file(path: str, selector=None, *args) -> dict:
        with open(path, encoding='utf-8') as file:
            data = yaml.safe_load(file)
            if selector:
                info = data.get(selector)
                if args:
                    return info.get(args[0])
                else:
                    return info
            else:
                return data

    # 写入yaml文件
    @staticmethod
    def write_yaml(file_name, content):
        with open(file_name, mode='w', encoding='utf-8') as file:
            yaml.dump(data=content, stream=file, allow_unicode=True)

    # 清空yaml文件内容
    @staticmethod
    def clearYaml(filename):
        with open(filename, mode='w', encoding='utf-8') as file:
            file.truncate()

    # 读取excel文件内容
    @staticmethod
    def read_excel(path, ws=None, columns=None) -> any:
        global worksheet
        try:
            workbook = openpyxl.load_workbook(path)
            if ws:
                try:
                    worksheet = workbook[ws]
                except Exception as e:
                    raise ValueError(f"{ws}工作表不存在，请检查一下。", e)
            else:
                worksheet = workbook.active
        except Exception as e:
            raise ValueError(f"{path}文件不存在或打开有误，有检查下文件.", e)
        # 获取工作表最大行数
        sheet_rows = worksheet.max_row
        print(f"{worksheet.title}工作薄的最大行数为:", sheet_rows)
        if sheet_rows == 0:
            raise ValueError(f"{worksheet.title}工作薄为空，请检查是否存在数据")
        else:
            # 获取真实行数
            real_sheet_rows = 0
            while sheet_rows > 0:
                row_dict = {i.value for i in worksheet[sheet_rows]}
                if row_dict == {None}:
                    sheet_rows -= 1
                else:
                    real_sheet_rows = sheet_rows
                    break
        # 初始化获取数据表第一行内容
        content_head = []
        for row in range(1, real_sheet_rows + 1):
            if row == 1:
                for col in range(1, worksheet.max_column + 1):
                    head = worksheet.cell(row=row, column=col).value
                    if head is None:
                        continue
                    else:
                        content_head.append(head)
        # 初始化空列表，用于保存获取表数据组的字典，#判断是否指定列参数，如指定，则获取指定列的内容。未指定时，获取全部表的内容。
        li = []
        if columns:
            for col in columns:
                # 获取第一行字段名
                head = worksheet[str(col) + '1'].value
                # 定义空列表，用于保存每列内容。
                col_content = []
                for i in range(2, real_sheet_rows + 1):
                    value = worksheet[str(col) + str(i)].value
                    col_content.append(value.strip())
                # dict[head]=col
                li.append(col_content)
                print(f"指定列{head}的内容数据长度为：", len(col_content))
            return li
        else:
            for row in range(2, real_sheet_rows + 1):
                content_li = []
                for col in range(1, len(content_head) + 1):
                    content = worksheet.cell(row=row, column=col).value
                    if content and type(content) is str:
                        content_li.append(content.strip())
                    else:
                        content_li.append(content)
                dict_obj = {}
                for length in range(len(content_head)):
                    dict_obj[content_head[length]] = content_li[length]
                li.append(dict_obj)
            print(f">>获取实际数据行数结果为：{len(li)}")
            return li


if __name__ == "__main__":
    # from config.setting import ConfigInfo
    # aa = ReadFile.read_yaml_file(ConfigInfo.ENV_CONFIG_FILE, "test")
    path_txt = r"D:\url.txt"
    content = ReadFile.read_txt_file(path_txt)
    res = []
    for i in content:
        link = i.strip("\n")
        value = link.split('/')[-1].split('.')[0]
        if value:
            res.append(int(value))
    print(res)
